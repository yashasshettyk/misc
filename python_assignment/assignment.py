import questions_collection as qc

ply_list = []
update = []
mx = 0
ind = 0
class game:
    def __init__(self,players_list):
        self.players = players_list

    def playersList(self):
        for j,i in enumerate(self.players):
            print(f"player {j+1} - {i} ")
    def return_player_name(self,i):
        return self.players[i]

class askquestions:
    def play(self):
        score = 0
        for i in range(0,len(qc.questions)):
            print(f"{i+1}) {qc.questions[i][0]}")
            for j,k in qc.questions[i][1].items():
                print(f"{j} : {k}")
            print()
            x = input("choose the right option : ")
            if x.upper() == qc.questions[i][2]:
                score += 1
        print(f"your total score is : {score}")
        update.append(score)


if __name__=="__main__":
    print("<-------------PYTHON QUIZZ------------->")
    ready = input("Are You Ready to Play the game ? (yes/no) : ")
    players_list = []
    total_players = 0
    if(ready.lower()=='yes'):
        total_players = int(input("Enter the total number of players : "))
        for i in range(total_players):
            players_list.append(input(f"Enter the name of player {i+1} : "))
    object = game(players_list)
    object.playersList()
    ply_list = players_list.copy()
    for i in range(total_players):

        plr = object.return_player_name(i)
        x = input(f"{plr} press enter to start")
        if x=="":
            plr = askquestions()
            print(f"\n----------------------------------------------------\n {object.return_player_name(i)}  playing \n----------------------------------------------------\n")
            total_score = plr.play()
            if (i+1)!=total_players:
                print(f"\n----------------------------------------------------")
                print("Next Player")
                print(f"----------------------------------------------------\n")


if __name__=="__main__":
    import openpyxl

    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active

    c1 = my_sheet.cell(row=1, column=1)
    c1.value = "Sl no."
    c1 = my_sheet.cell(row=1, column=2)
    c1.value = "players"
    c1 = my_sheet.cell(row=1, column=3)
    c1.value = "points"
    c1 = my_sheet.cell(row=1, column=4)
    c1.value = "marks"

    for count,names in enumerate(ply_list):
        c1 = my_sheet.cell(row=count+2, column=1)
        c1.value = count+1
        c1 = my_sheet.cell(row=count + 2, column=2)
        c1.value = names
        c1 = my_sheet.cell(row=count + 2, column=3)
        c1.value = update[count]

        mark = (update[count] / len(qc.questions)) * 100

        c1 = my_sheet.cell(row=count + 2, column=4)
        c1.value = mark
        my_wb.save("score.xlsx")





















